import datetime
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

# set up the Chrome driver
driver = webdriver.Chrome()

# load the input file
input_wb = openpyxl.load_workbook('D://Python//Automation//input.xlsx')
input_ws = input_wb.active

# create the output workbook and sheet
output_wb = openpyxl.Workbook()
output_ws = output_wb.active
output_ws.title = 'Output'
output_ws['A1'] = 'Vendor'
output_ws['B1'] = 'Lastname/Firstname'
output_ws['C1'] = 'Status'
output_ws['D1'] = 'Execution Date'

# loop through each row in the input file
for row in input_ws.iter_rows(min_row=2, values_only=True):
    vendor = row[0]
    lastname_firstname = row[1]
    firstname, lastname = lastname_firstname.split(' ', 1)
    lastname = lastname.replace(",", "")  # Remove comma from last name

    # navigate to the OIG Exclusions website
    driver.get('https://exclusions.oig.hhs.gov/')

    # wait for the Last Name field to be visible and enter the last name
    last_name_box = WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.NAME, "ctl00$cpExclusions$txtSPLastName")))
    last_name_box.clear()
    last_name_box.send_keys(lastname)

    # wait for the First Name field to be visible and enter the first name
    first_name_box = WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.NAME, "ctl00$cpExclusions$txtSPLastName")))
    first_name_box.clear()
    first_name_box.send_keys(firstname)

    # click the Search button
    search_button = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.NAME, "ctl00$cpExclusions$ibSearchSP")))
    search_button.click()

    # wait for the results table to load and check if there are any results
    no_results_msg = 'No records found.'
    try:
        WebDriverWait(driver, 50).until(EC.text_to_be_present_in_element((By.XPATH, "//table[@id='ctl00_ContentPlaceHolder1_grdExclusions']/tbody/tr[2]/td"), no_results_msg))
        status = 'Clear'
    except:
        status = 'Not Clear'

    # add the row to the output file
    output_ws.append([vendor, lastname_firstname, status, datetime.datetime.now()])

    # take a screenshot and save it
    now = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    driver.save_screenshot(f'D://Python//Automation//Screenshots//{vendor}_{now}.png')

# save the output file and close the Chrome driver
output_wb.save('D://Python//Automation//output.xlsx')
driver.quit()
